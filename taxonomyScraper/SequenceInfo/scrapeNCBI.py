from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import sys
from openpyxl import *
from openpyxl.styles import Font, Border, Side
from selenium import webdriver
import time
import gff3Counter as count

url = 'https://www.ncbi.nlm.nih.gov/nuccore/'

from selenium.webdriver.chrome.options import Options

def getSeq(url):
    options = webdriver.ChromeOptions()
    prefs = {"download.default_directory" : "/home/nate/repos/BioScripts/taxonomyScraper"}
    options.add_experimental_option("prefs",prefs)
    driver = webdriver.Chrome(chrome_options=options)

    driver.get(url)
    # time.sleep(5)
    # driver.find_element_by_xpath('html/body/div[1]/a[2]').click()
    # element = WebDriverWait(driver, secs).until(find)
    time.sleep(2)
    driver.find_element_by_xpath(".//*[@id='seqsendto']/a").click()
    time.sleep(1)
    driver.find_element_by_xpath(".//*[@id='dest_File']").click()
    time.sleep(1)
    driver.find_element_by_xpath(".//*[@id='file_format']").click()
    time.sleep(.3)
    driver.find_element_by_xpath(".//*[@id='file_format']/option[12]").click()
    time.sleep(.4)
    driver.find_element_by_xpath(".//*[@id='submenu_File']/button").click()
    time.sleep(5)
    driver.close()

    
def writeDic(ws,index,dic,acces):
    ws.cell(row=index,column=13).font = Font(name='Calibri')
    ws.cell(row=index,column=13).value = acces
    col = 14
    for key in dic[acces].keys():
        ws.cell(row=index,column=col).font = Font(name='Calibri')
        ws.cell(row=index,column=col).value = dic[acces][key]
        col+=1


#Process the file and then wait
import pprint
import os
import re
from openpyxl import *
from openpyxl.styles import Font, Border, Side
accre = re.compile('Accession',re.IGNORECASE)

title = sys.argv[1]
wb = load_workbook(str(sys.argv[1]))
wb.guess_types = True
ws = wb.worksheets[0]
for index,row in enumerate(ws.rows,start=1):
    if index != 1:
        acces = row[8].value
        match = accre.search(str(acces))
        if acces != None and not match:
            try:
                getSeq(url+acces)
                dic = count.counter('sequence.gff3',acces)
                writeDic(ws,index,dic,acces)
                os.remove('sequence.gff3')
            except:
                ws.cell(row=index,column=13).font = Font(name='Calibri')
                ws.cell(row=index,column=13).value = acces
                os.remove('sequence.gff3')
    else:
        ws.cell(row=index,column=13).font = Font(name='Calibri')
        ws.cell(row=index,column=13).value = "Accession"
        ws.cell(row=index,column=14).font = Font(name='Calibri')
        ws.cell(row=index,column=14).value = "CDS"
        ws.cell(row=index,column=15).font = Font(name='Calibri')
        ws.cell(row=index,column=15).value = "rRNA"
        ws.cell(row=index,column=16).font = Font(name='Calibri')
        ws.cell(row=index,column=16).value = "rRNA"

# wb.save('plastid1.xlsx')        







# Subclass Order Taxa Size(bp) GCContent(%) CDSa tRNA rRNA Accession Reference Citation	Website
# I'm going to try to respect ncbi and use their interface but I might need to scrape for some things.
import os
import re
from openpyxl import *
from openpyxl.styles import Font, Border, Side,Alignment
accre = re.compile('Accession',re.IGNORECASE)
from Bio.SeqUtils import GC
from Bio import Entrez as E
from Bio import SeqIO

def CountFeatures(r):
    fcount = {}
    for f in r.features:
        fcount[f.type] = 0 
    for f in r.features:
        fcount[f.type] +=1
    return(fcount)


def WriteFeatures(r,ws,index):
    print(r.name)
    fdic = CountFeatures(r)
    print(fdic)
    if 'CDS' in fdic.keys() and 'tRNA' in fdic.keys() and 'rRNA' in fdic.keys():
        features = [fdic['CDS'], fdic['tRNA'],fdic['rRNA']]
        col = 6
        featNum = 0
        for key in features:
            ws.cell(row=index,column=col).alignment = Alignment(horizontal="center")
            ws.cell(row=index,column=col).font = Font(name='Times New Roman')
            ws.cell(row=index,column=col).value = features[featNum]
            col+=1
            featNum+=1


def CheckTaxaNames(ws,seq_dic):
    for index,row in enumerate(ws.rows,start=1):
        #don't think I need this condition
        if index != 1:
            acces = row[8].value
            accre = re.compile('Accession',re.IGNORECASE)
            match = accre.search(str(acces))
            if acces != None and not match:
                acces = acces.replace('\xa0','')
                nameRe = re.compile(str(row[2].value),re.IGNORECASE)
                genSpec = str(row[2].value).split()
                if len(genSpec) < 2:
                    genSpec.append('')
                    genusRe = re.compile(genSpec[0],re.IGNORECASE)
                    speciesRe = re.compile(genSpec[1],re.IGNORECASE)
                    org = seq_dic[acces].annotations['organism']
                    if not nameRe.search(org):
                        if genusRe.search(org) or speciesRe.search(org) or row[2] == None:
                            print('spelling, ',row[2].value,acces)
                            ws.cell(row=index,column=3).alignment = Alignment(horizontal="center")
                            ws.cell(row=index,column=3).font = Font(name='Times New Roman',italic=True,size=12)
                            ws.cell(row=index,column=3).value = org

                        else:
                            print(row[2].value,acces)

def FillOrder(r,ws,index):
    ws.cell(row=index,column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=index,column=2).font = Font(name='Times New Roman',size=12)
    ws.cell(row=index,column=2).value = r.annotations['taxonomy'][-3]

def FillAllOrder(seq_dic,ws):
    accre = re.compile('Accession',re.IGNORECASE)
    for index,row in enumerate(ws.rows,start=1):
        acces = row[8].value
        match = accre.search(str(acces))
        if acces != None and not match and row[1].value == None:
            acces = acces.replace('\xa0','')
            FillOrder(seq_dic[acces],ws,index)


def CheckSizes(seq_dic,ws):
    accre = re.compile('Accession',re.IGNORECASE)
    for index,row in enumerate(ws.rows,start=1):
        acces = row[8].value
        match = accre.search(str(acces))
        if acces != None and not match:
            acces = acces.replace('\xa0','')
            seqL = len(seq_dic[acces].seq)
            excelL = int(str(row[3].value).replace(',',''))
            if seqL != excelL:
                ws.cell(row=index,column=4).alignment = Alignment(horizontal="center")
                ws.cell(row=index,column=4).font = Font(name='Times New Roman',size=12)
                ws.cell(row=index,column=4).value = seqL

    

def FillFeatures(seq_dic,ws):
    accre = re.compile('Accession',re.IGNORECASE)
    for index,row in enumerate(ws.rows,start=1):
        acces = row[8].value
        match = accre.search(str(acces))
        if acces != None and not match:
            acces = acces.replace('\xa0','')
            WriteFeatures(seq_dic[acces],ws,index)


def FillGC(seq_dic,ws):
    accre = re.compile('Accession',re.IGNORECASE)
    for index,row in enumerate(ws.rows,start=1):
        acces = row[8].value
        match = accre.search(str(acces))
        if acces != None and not match:
            acces = acces.replace('\xa0','')
            ws.cell(row=index,column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=index,column=5).font = Font(name='Times New Roman',size=12)
            ws.cell(row=index,column=5).value = round(GC(seq_dic[acces].seq),2)
            



r = SeqIO.parse("seqs.gb","genbank")
seq_dic = SeqIO.to_dict(r,key_function=lambda x: x.name)

wb = load_workbook('p-mito4-3-17NewFeatures.xlsx')
wb.guess_types = True
ws = wb.worksheets[0]

FillGC(seq_dic,ws)

wb.save('p-mito4-3-17NewFeatures.xlsx')

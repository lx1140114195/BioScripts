# Lets not scrape their site since they have an interface for it
import os
import re
from openpyxl import *
from openpyxl.styles import Font, Border, Side
accre = re.compile('Accession',re.IGNORECASE)
# wb = load_workbook(str(sys.argv[1]))
# wb.guess_types = True
# ws = wb.worksheets[0]
from Bio import Entrez
Entrez.email = "ns196414@ohio.edu"     # Always tell NCBI who you are
handle = Entrez.efetch(db="nucleotide", id="NC_020795", rettype="gb", retmode="text")
print(handle.read())



# for index,row in enumerate(ws.rows,start=1):
#     if index != 1:
#         acces = row[8].value
#         match = accre.search(str(acces))
#         if acces != None and not match:
#             try:
#                 getSeq(url+acces)
#                 dic = count.counter('sequence.gff3',acces)
#                 writeDic(ws,index,dic,acces)
#                 os.remove('sequence.gff3')
#             except:
#                 ws.cell(row=index,column=13).font = Font(name='Calibri')
#                 ws.cell(row=index,column=13).value = acces
#                 os.remove('sequence.gff3')
#     else:
#         ws.cell(row=index,column=13).font = Font(name='Calibri')
#         ws.cell(row=index,column=13).value = "Accession"
#         ws.cell(row=index,column=14).font = Font(name='Calibri')
#         ws.cell(row=index,column=14).value = "CDS"
#         ws.cell(row=index,column=15).font = Font(name='Calibri')
#         ws.cell(row=index,column=15).value = "rRNA"
#         ws.cell(row=index,column=16).font = Font(name='Calibri')
#         ws.cell(row=index,column=16).value = "rRNA"

# wb.save('TableFromAcc.xlsx')        

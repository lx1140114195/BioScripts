from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import sys
from openpyxl import *
from openpyxl.styles import Font, Border, Side

from selenium import webdriver
import time

def find(driver):
    element = driver.find_elements_by_id("data")
    if element:
        return element
    else:
        return False

out = open('out.txt','w')
f = open('noResults.txt','w')
def getTaxa(url,genus):
    driver = webdriver.PhantomJS()
    driver.get(url)
  
    # time.sleep(5)
    # driver.find_element_by_xpath('html/body/div[1]/a[2]').click()
    # element = WebDriverWait(driver, secs).until(find)
    time.sleep(2)
    search = driver.find_element_by_id('gettaxon')
    search.send_keys(genus)
    time.sleep(1)
    driver.execute_script("window.scrollTo(0, 20)")
    time.sleep(1)
    search.submit()

    subclass = ''
    order = ''
    if len(driver.find_elements_by_id('gettaxon')) > 0:
        f.write(genus)
    else:
        time.sleep(1)
        try:
            subclass = driver.find_element_by_xpath(".//*[@id='contentright']/table/tbody/tr[15]/td[2]/p/a").text
            order = driver.find_element_by_xpath(".//*[@id='contentright']/table/tbody/tr[17]/td[2]/p/a").text
        except:
            f.write(genus)
    driver.delete_all_cookies()
    driver.close()
    return([subclass,order])

# sub = driver.find_element_by_name('Submit')
# sub.click()


def emptyOrMatch(r,i):
    R = r.search(i)
    if str(type(R)) != "<class 'NoneType'>":
        return R.group(0)
    else:
        return ''

def grabFirstColumn(s):
    l = []
    for row in s.iter_rows('A{}:A{}'.format(s.min_row,s.max_row)):
        for cell in row:
            l.append(cell.value)
    return l

def sil(x):
    if x is None:
        return('')
    else:
        return(str(x))

def populate(col,columnList):
    for i in range(1,len(columnList)+1):
        worksheet.write(i,col,columnList[i-1])

#Scraping for the taxonomy of algae in an excel sheet
url = "http://www.algaebase.org/browse/taxonomy/"
title = sys.argv[1]

wb = load_workbook(str(sys.argv[1]))
wb.guess_types = True
ws = wb.worksheets[0]
#fixes some stuff lol third column adds std or (Mi|So)

import re
taxare = re.compile('Taxa',re.IGNORECASE)


for index,row in enumerate(ws.rows,start=1):
    species = row[2].value
    match = taxare.search(str(species))
    if species != None and not match:
        genus = species.split()
        tax = getTaxa(url,genus[0])
        ws.cell(row=index,column=1).font = Font(name='Calibri')
        ws.cell(row=index,column=1).value = tax[0]
        ws.cell(row=index,column=2).font = Font(name='Calibri')
        ws.cell(row=index,column=2).value = tax[1]


ws.save('plastid2.xlsx')

        # print(tax[0],tax[1],file=out)


    
    #Still need to do some writing for this excel file
    # ws.cell(row=index,column=3).value = ma.group(0)
    # ws.cell(row=index,column=3).font = Font(name='Calibri')


# ws.cell(row=1,column=18).value = ''

out.close()
f.close()
# subclass, order = getTaxa(url,genus)
# print(subclass,order)




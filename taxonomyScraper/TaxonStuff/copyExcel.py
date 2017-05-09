from openpyxl import *
from openpyxl.styles import Alignment
import sys
from openpyxl.styles import Font, Border, Side
import re
accre = re.compile('Taxa',re.IGNORECASE)

wb = load_workbook(str(sys.argv[1]))
wb.guess_types = True
ws = wb.worksheets[0]
taxa = []
f = open('out.txt','r')

for line in f:
    lineL = line.split()
    taxa.append(lineL)
count = 0
with open('taxaOutput.txt','w') as ta:
    for index,row in enumerate(ws.rows,start=1):
        acces = row[2].value
        match = accre.search(str(acces))
        if index != 1 and not match and acces != None and row[8].value != None:
            print(row[8].value,end=' ',file=ta)
            ws.cell(row=index,column=1).font = Font(name='Times New Roman',size=12)
            ws.cell(row=index,column=2).font = Font(name='Times New Roman',size=12)
            if len(taxa[count]) < 3:
                print()
                pass
            else:
                l = taxa[count]
                print(l[1],l[2],file=ta)
                ws.cell(row=index,column=1).value = l[1]
                ws.cell(row=index,column=2).value = l[2]
                ws.cell(row=index,column=2).alignment = Alignment(horizontal="center")
                ws.cell(row=index,column=1).alignment = Alignment(horizontal="center")

            count+=1



wb.save('plastidEdit1.xlsx')        
